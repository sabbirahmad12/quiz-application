import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import load_workbook
from excel_db import get_all_quizzes, get_leaderboard_data, add_new_quiz, add_question_to_quiz
from excel_db import (
    QUIZZES_FILE,
    QUESTIONS_FILE,
    get_all_quizzes,
)

class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip_window = None
        self.widget.bind("<Enter>", self.show_tip)
        self.widget.bind("<Leave>", self.hide_tip)

    def show_tip(self, event=None):
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 25
        
        self.tip_window = tk.Toplevel(self.widget)
        self.tip_window.wm_overrideredirect(True)
        self.tip_window.wm_geometry(f"+{x}+{y}")
        
        label = tk.Label(
            self.tip_window, 
            text=self.text, 
            background="#ffffe0", 
            relief="solid", 
            borderwidth=1
        )
        label.pack()

    def hide_tip(self, event=None):
        if self.tip_window:
            self.tip_window.destroy()
            self.tip_window = None

class TeacherDashboard(tk.Toplevel):
    def __init__(self, master, user_id, user_name=None):
        super().__init__(master)
        self.title("Teacher Dashboard")
        self.geometry("1200x800")
        self.user_id = user_id
        self.user_name = user_name

        # Main frame
        self.main_frame = ttk.Frame(self)
        self.main_frame.pack(fill='both', expand=True, padx=20, pady=20)

        # Initialize widgets
        self._init_widgets()

        # Set up the UI
        self._setup_ui()

        # Load initial data
        self._load_initial_data()

        # Make the window visible
        self.deiconify()

    def _init_widgets(self):
        """Initialize all UI components."""
        self.notebook = None
        self.quiz_tree = None
        self.student_tree = None
        self.leaderboard_tree = None
        self.correct_option = None
        self._error_label = None

    def _setup_ui(self):
        """Set up the main UI layout."""
        self.notebook = ttk.Notebook(self.main_frame)  # Attach to main_frame
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)

        # Create tab frames
        self.quiz_tab = ttk.Frame(self.notebook)
        self.students_tab = ttk.Frame(self.notebook)
        self.leaderboard_tab = ttk.Frame(self.notebook)
        self.add_quiz_tab = ttk.Frame(self.notebook)

        # Add tabs to the notebook
        self.notebook.add(self.quiz_tab, text="Manage Quizzes")
        self.notebook.add(self.students_tab, text="Students")
        self.notebook.add(self.leaderboard_tab, text="Leaderboard")
        self.notebook.add(self.add_quiz_tab, text="Add New Quiz")

        # Set up each tab
        self._setup_quiz_tab()
        self._setup_students_tab()
        self._setup_leaderboard_tab()
        self._setup_add_quiz_tab()

    def _load_initial_data(self):
        """Load initial data into the UI."""
        print("Loading quizzes...")
        self._load_quizzes()
        print("Loading students...")
        self._load_students()
        print("Loading leaderboard...")
        self._load_leaderboard()

    def _setup_quiz_tab(self):
        # Quiz Management UI
        ttk.Label(self.quiz_tab, text="All Quizzes", style='Title.TLabel').pack(pady=10)

        self.quiz_tree = ttk.Treeview(self.quiz_tab, 
                                    columns=('ID', 'Subject', 'Title', 'Description'), 
                                    show='headings')
        self.quiz_tree.heading('ID', text='ID')
        self.quiz_tree.heading('Subject', text='Subject')
        self.quiz_tree.heading('Title', text='Title')
        self.quiz_tree.heading('Description', text='Description')
        
        self.quiz_tree.column('ID', width=50, anchor='center')
        self.quiz_tree.column('Subject', width=150)
        self.quiz_tree.column('Title', width=200)
        self.quiz_tree.column('Description', width=300)
        
        scrollbar = ttk.Scrollbar(self.quiz_tab, orient="vertical", command=self.quiz_tree.yview)
        self.quiz_tree.configure(yscrollcommand=scrollbar.set)
        
        self.quiz_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        btn_frame = ttk.Frame(self.quiz_tab)
        btn_frame.pack(pady=10)
        
        ttk.Button(btn_frame, text="Refresh", command=self._load_quizzes).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Delete Selected", command=self._delete_quiz).pack(side='left', padx=5)
    
        filter_frame = ttk.Frame(self.quiz_tab)
        filter_frame.pack(fill='x', pady=5)
        
        ttk.Label(filter_frame, text="Filter by Subject:").pack(side='left', padx=5)
        self.subject_filter = ttk.Combobox(filter_frame, values=["All", "Mathematics", "Bangla", "English", "General Knowledge"])
        self.subject_filter.pack(side='left', padx=5)
        self.subject_filter.set("All")
        
        ttk.Button(filter_frame, text="Apply Filter", command=self._apply_filter).pack(side='left', padx=5)

    def _apply_filter(self):
        selected_subject = self.subject_filter.get()
        for item in self.quiz_tree.get_children():
            self.quiz_tree.delete(item)
        
        quizzes = get_all_quizzes()
        for quiz in quizzes:
            # Extract subject from the title
            quiz_subject = quiz['title'].split(':')[0].strip() if ':' in quiz['title'] else "General"
            
            if selected_subject == "All" or quiz_subject == selected_subject:
                # Extract subject and title
                if ':' in quiz['title']:
                    subject, title = quiz['title'].split(':', 1)
                    subject = subject.strip()
                    title = title.strip()
                else:
                    subject = "General"
                    title = quiz['title']
                
                self.quiz_tree.insert('', 'end', 
                                    values=(quiz['id'], 
                                        subject,
                                        title,
                                        quiz.get('description', '')))


    def _setup_students_tab(self):
        # Student Management UI
        ttk.Label(self.students_tab, text="Student List", style='Title.TLabel').pack(pady=10)
        
        self.student_tree = ttk.Treeview(self.students_tab, columns=('ID', 'Username'), show='headings')
        self.student_tree.heading('ID', text='ID')
        self.student_tree.heading('Username', text='Username')
        self.student_tree.pack(fill='both', expand=True)

        btn_frame = ttk.Frame(self.students_tab)
        btn_frame.pack(pady=10)
        
        ttk.Button(btn_frame, 
                text="Refresh List", 
                command=self._load_students).pack(side='left', padx=5)

    def _setup_leaderboard_tab(self):
        # Leaderboard UI
        ttk.Label(self.leaderboard_tab, text="Leaderboard", style='Title.TLabel').pack(pady=10)
        
        self.leaderboard_tree = ttk.Treeview(self.leaderboard_tab, 
                                           columns=('Rank', 'Student', 'Quiz', 'Score'), 
                                           show='headings')
        self.leaderboard_tree.heading('Rank', text='Rank')
        self.leaderboard_tree.heading('Student', text='Student')
        self.leaderboard_tree.heading('Quiz', text='Quiz')
        self.leaderboard_tree.heading('Score', text='Score')
        self.leaderboard_tree.pack(fill='both', expand=True)

        # Add refresh button at the bottom
        btn_frame = ttk.Frame(self.leaderboard_tab)
        btn_frame.pack(pady=10)
        
        ttk.Button(btn_frame, 
                text="Refresh Leaderboard", 
                command=self._load_leaderboard).pack(ipadx=20, pady=5)

    def _setup_add_quiz_tab(self):
        # Main Frame
        main_frame = ttk.Frame(self.add_quiz_tab)
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Quiz Metadata
        ttk.Label(main_frame, text="Subject/Category:").grid(row=0, column=0, sticky='e', pady=5)
        self.quiz_category = ttk.Combobox(main_frame, values=["Mathematics", "Bangla", "English", "General Knowledge"])
        self.quiz_category.grid(row=0, column=1, sticky='ew', pady=5)
        
        ttk.Label(main_frame, text="Quiz Title:").grid(row=1, column=0, sticky='e', pady=5)
        self.quiz_title_entry = ttk.Entry(main_frame)
        self.quiz_title_entry.grid(row=1, column=1, sticky='ew', pady=5)
        
        ttk.Label(main_frame, text="Description:").grid(row=2, column=0, sticky='e', pady=5)
        self.quiz_desc_entry = ttk.Entry(main_frame)
        self.quiz_desc_entry.grid(row=2, column=1, sticky='ew', pady=5)
        
        ttk.Separator(main_frame).grid(row=3, column=0, columnspan=2, pady=10, sticky='ew')
        ttk.Label(main_frame, text="Add Questions", style='Heading.TLabel').grid(row=4, column=0, columnspan=2)
        
        self.questions_frame = ttk.Frame(main_frame)
        self.questions_frame.grid(row=5, column=0, columnspan=2, sticky='nsew')
        
        self.questions_tree = ttk.Treeview(self.questions_frame, columns=('Question', 'Options', 'Answer'), show='headings', height=5)
        self.questions_tree.heading('Question', text='Question')
        self.questions_tree.heading('Options', text='Options')
        self.questions_tree.heading('Answer', text='Correct Answer')
        self.questions_tree.pack(side='left', fill='both', expand=True)
        
        # Question Add Form
        form_frame = ttk.Frame(self.questions_frame)
        form_frame.pack(side='right', fill='y', padx=10)
        
        ttk.Label(form_frame, text="Question:").pack(anchor='w')
        self.question_entry = ttk.Entry(form_frame, width=30)
        self.question_entry.pack(fill='x', pady=5)
        
        ttk.Label(form_frame, text="Options:").pack(anchor='w')
        self.option1 = ttk.Entry(form_frame, width=30)
        self.option1.pack(fill='x', pady=2)
        self.option2 = ttk.Entry(form_frame, width=30)
        self.option2.pack(fill='x', pady=2)
        self.option3 = ttk.Entry(form_frame, width=30)
        self.option3.pack(fill='x', pady=2)
        self.option4 = ttk.Entry(form_frame, width=30)
        self.option4.pack(fill='x', pady=2)
        
        ttk.Label(form_frame, text="Correct Option (1-4):").pack(anchor='w')
        self.correct_option = ttk.Spinbox(
            form_frame, 
            from_=1, 
            to=4, 
            width=5,
            validate="key",
            validatecommand=(self.register(self._validate_correct_option), '%P')
        )
        self.correct_option.pack(anchor='w', pady=5)
        
        ToolTip(self.correct_option, "Enter: 1=Option1, 2=Option2, 3=Option3, 4=Option4")
        
        ttk.Button(form_frame, text="Add Question", command=self._add_question).pack(pady=10)
        ttk.Button(form_frame, text="Remove Selected", command=self._remove_question).pack(pady=5)

        ttk.Button(main_frame, text="Save Quiz", command=self._save_quiz_with_questions, style='Accent.TButton').grid(row=6, column=0, columnspan=2, pady=20)

        self.correct_option.bind('<FocusOut>', self._check_correct_option)
        self.correct_option.bind('<Return>', lambda e: self._add_question())

    def _reset_quiz_form(self):
        self.quiz_category.set('')
        self.quiz_title_entry.delete(0, 'end')
        self.quiz_desc_entry.delete(0, 'end')
        for item in self.questions_tree.get_children():
            self.questions_tree.delete(item)

    def _check_correct_option(self, event=None):
        value = self.correct_option.get()
        if not value:
            return
            
        try:
            num = int(value)
            if not 1 <= num <= 4:
                self._show_input_error("Must be between 1-4")
        except ValueError:
            self._show_input_error("Numbers only (1-4)")

    def _validate_correct_option(self, new_value):
        if new_value == "":
            return True
        try:
            value = int(new_value)
            return 1 <= value <= 4
        except ValueError:
            return False

    def _load_quizzes(self):
        for item in self.quiz_tree.get_children():
            self.quiz_tree.delete(item)
        
        quizzes = get_all_quizzes()
        if not quizzes:
            messagebox.showinfo("Info", "No quizzes available.")
            return
        
        for i, quiz in enumerate(quizzes, 1):  # Start from 1
            if ':' in quiz['title']:
                subject, title = quiz['title'].split(':', 1)
                subject = subject.strip()
                title = title.strip()
            else:
                subject = "General"
                title = quiz['title']
            
            self.quiz_tree.insert('', 'end', 
                                values=(i,  # Use i instead of quiz['id']
                                    subject,
                                    title,
                                    quiz.get('description', '')))
    
    def _load_students(self):
        self.student_tree.delete(*self.student_tree.get_children())
        
        try:
            from excel_db import get_all_students
            students = get_all_students()
            
            if not students:
                messagebox.showinfo("Info", "No students registered yet.")
                return
                
            # Check The Unique ID and username
            seen = set()
            for i, student in enumerate(students, 1): 
                identifier = (student['id'], student['username'])
                if identifier not in seen:
                    self.student_tree.insert(
                        '', 'end', 
                        values=(i, student['username']) 
                    )
                    seen.add(identifier)
                    
        except Exception as e:
            messagebox.showerror(
                "Database Error", 
                f"Failed to load students: {str(e)}"
            )
    
    def _load_leaderboard(self):
        for item in self.leaderboard_tree.get_children():
            self.leaderboard_tree.delete(item)
        
        leaderboard = get_leaderboard_data()
        for i, entry in enumerate(leaderboard[:20], 1):
            self.leaderboard_tree.insert('', 'end', 
                                       values=(i, entry['student_name'], 
                                              entry['quiz_title'], entry['score']))
            
    def _add_question(self):
        """Add a new question to the quiz."""
        try:
            # Validate inputs
            question = self.question_entry.get().strip()
            options = [
                self.option1.get().strip(),
                self.option2.get().strip(),
                self.option3.get().strip(),
                self.option4.get().strip()
            ]
            
            # Check for empty fields
            if not question or any(not opt for opt in options):
                raise ValueError("All fields must be filled")

            # Get and validate correct option index (1-4)
            correct_index = int(self.correct_option.get())
            if correct_index < 1 or correct_index > 4:
                raise ValueError("Correct option must be 1-4")

            # Format options for display (1-4 numbering)
            options_display = [f"{i+1}. {opt}" for i, opt in enumerate(options)]
            
            # Add to treeview
            self.questions_tree.insert('', 'end', values=(
                question, 
                "\n".join(options_display),
                f"{correct_index}. {options[correct_index-1]}"
            ))

            # Clear form
            self._clear_question_form()
            messagebox.showinfo("Success", "Question added!")
            
        except ValueError as e:
            self._show_input_error(str(e))
        except Exception as e:
            self._show_input_error(f"Unexpected error: {str(e)}")

    def _show_input_error(self, message):
        if not self._error_label:
            self._error_label = ttk.Label(
                self.add_quiz_tab,
                text="",
                foreground='red',
                font=('Arial', 10)
            )
            self._error_label.pack(pady=5)

        self._error_label.config(text=message)

        # Automatically hide the error message after 3 seconds
        self.after(3000, lambda: self._error_label.pack_forget())

    def _clear_question_form(self):
        self.question_entry.delete(0, tk.END)
        for option in [self.option1, self.option2, self.option3, self.option4]:
            option.delete(0, tk.END)
        self.correct_option.delete(0, tk.END)
        self.correct_option.insert(0, "0")

    def _remove_question(self):
        selected_item = self.questions_tree.selection()
        if not selected_item:
            messagebox.showwarning("Warning", "Please select a question to remove!")
            return

        # Remove the selected item
        for item in selected_item:
            self.questions_tree.delete(item)

    def _save_quiz_with_questions(self):
        try:
            # Validate quiz info
            if not self.quiz_category.get() or not self.quiz_title_entry.get():
                raise ValueError("Category and Title are required")
                
            if not self.questions_tree.get_children():
                raise ValueError("Add at least one question")

            # Save quiz
            quiz_id = add_new_quiz(
                f"{self.quiz_category.get()}: {self.quiz_title_entry.get()}",
                self.quiz_desc_entry.get()
            )

            # Process each question
            for item in self.questions_tree.get_children():
                question, options_text, correct_answer = self.questions_tree.item(item)['values']
                
                # Extract options (remove numbering)
                options = [opt.split('. ', 1)[1] for opt in options_text.split('\n')]
                
                correct_num = int(correct_answer.split('.', 1)[0])
                correct_index = correct_num - 1 
                
                # Save to database
                add_question_to_quiz(quiz_id, question, options, correct_index)

            # Reset form on success
            self._reset_quiz_form()
            messagebox.showinfo("Success", "Quiz saved successfully!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save: {str(e)}")

    def _delete_quiz(self):
    
        selected_item = self.quiz_tree.focus()
        if not selected_item:
            messagebox.showwarning("Warning", "Please select a quiz first!")
            return
        
        item_data = self.quiz_tree.item(selected_item)
        quiz_id, quiz_title = item_data['values'][0], item_data['values'][1]

        confirm = messagebox.askyesno(
            "Confirm Deletion",
            f"Are you sure you want to delete this quiz?\n\nTitle: {quiz_title}\nID: {quiz_id}",
            icon='warning'
        )
        
        if not confirm:
            return
        
        try:
            wb = load_workbook(QUIZZES_FILE)
            ws = wb.active

            for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if row[0].value == quiz_id: 
                    ws.delete_rows(idx)
                    break

            if QUESTIONS_FILE.exists():
                wb_q = load_workbook(QUESTIONS_FILE)
                ws_q = wb_q.active
                
                for idx in range(ws_q.max_row, 1, -1):
                    if ws_q.cell(row=idx, column=2).value == quiz_id:
                        ws_q.delete_rows(idx)
                
                wb_q.save(QUESTIONS_FILE)

            wb.save(QUIZZES_FILE)

            self._load_quizzes()
            messagebox.showinfo("Success", "Quiz deleted successfully!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete quiz: {str(e)}")
