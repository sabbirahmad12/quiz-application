import os
from openpyxl import Workbook, load_workbook
from pathlib import Path
from datetime import datetime

DB_FOLDER = Path(__file__).parent / "data"
DB_FOLDER.mkdir(exist_ok=True)  

USERS_FILE = DB_FOLDER / "users.xlsx"
QUIZZES_FILE = DB_FOLDER / "quizzes.xlsx"
QUESTIONS_FILE = DB_FOLDER / "questions.xlsx"
LEADERBOARD_FILE = DB_FOLDER / "leaderboard.xlsx"

def init_excel_db():
    print("Initializing Excel database...")
    if not USERS_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(["id", "username", "password", "role"])
        wb.save(USERS_FILE)
    
    if not QUIZZES_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(["id", "title", "description"])
        wb.save(QUIZZES_FILE)
    
    if not QUESTIONS_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(["id", "quiz_id", "question_text", "option1", "option2", "option3", "option4", "correct_answer"])
        wb.save(QUESTIONS_FILE)
    
    if not LEADERBOARD_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(["id", "user_id", "quiz_id", "score", "time_taken"])
        wb.save(LEADERBOARD_FILE)

def add_user(username, password, role):
    wb = load_workbook(USERS_FILE)
    ws = wb.active
    new_id = len(ws['A'])
    ws.append([new_id, username, password, role])
    wb.save(USERS_FILE)
    return new_id

def get_user(username):
    wb = load_workbook(USERS_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == username:
            return {"id": row[0], "username": row[1], "password": row[2], "role": row[3]}
    return None

def add_quiz(title, description=""):
    wb = load_workbook(QUIZZES_FILE)
    ws = wb.active
    new_id = len(ws['A'])
    ws.append([new_id, title, description])
    wb.save(QUIZZES_FILE)
    return new_id

def get_all_quizzes():
    wb = load_workbook(QUIZZES_FILE)
    ws = wb.active
    return [{"id": row[0], "title": row[1], "description": row[2]} 
            for row in ws.iter_rows(min_row=2, values_only=True)]

def add_question_to_quiz(quiz_id, question_text, options, correct_index):
    wb = load_workbook(QUESTIONS_FILE)
    ws = wb.active
    
    new_id = len(ws['A']) + 1
    ws.append([
        new_id,
        quiz_id,
        question_text,
        options[0],
        options[1],
        options[2],
        options[3],
        correct_index
    ])
    
    wb.save(QUESTIONS_FILE)

def get_quiz_questions(quiz_id):
    wb = load_workbook(QUESTIONS_FILE)
    ws = wb.active
    questions = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == quiz_id:
            questions.append({
                'id': row[0],
                'quiz_id': row[1],
                'question_text': row[2],
                'options': [row[3], row[4], row[5], row[6]],
                'correct_answer': row[7]
            })
    return questions

def save_score(user_id, user_name, quiz_id, score, time_taken):
    if not user_name or user_name.startswith("Student_"):
        user_name = (user_id) or f"Student_{user_id}"
    wb = load_workbook(LEADERBOARD_FILE)
    ws = wb.active
    new_id = len(ws['A'])
    ws.append([new_id, user_id, quiz_id, score, time_taken])
    wb.save(LEADERBOARD_FILE)

def get_top_scores(quiz_id=None, limit=10):
    wb = load_workbook(LEADERBOARD_FILE)
    ws = wb.active
    scores = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if quiz_id is None or row[2] == quiz_id:
            scores.append(row)
    
    scores.sort(key=lambda x: x[3], reverse=True)
    return scores[:limit]

def get_all_quizzes():
    if not QUIZZES_FILE.exists():
        return []
    
    wb = load_workbook(QUIZZES_FILE)
    ws = wb.active
    return [{"id": row[0], "title": row[1], "description": row[2]} 
            for row in ws.iter_rows(min_row=2, values_only=True)]

def get_all_students():
    if not USERS_FILE.exists():
        return []
    
    wb = load_workbook(USERS_FILE)
    ws = wb.active
    
    students = []
    seen_usernames = set()
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 4 and row[3] == "student" and row[1] not in seen_usernames:
            students.append({"id": row[0], "username": row[1]})
            seen_usernames.add(row[1])
    
    return students

def get_leaderboard_data():
    if not LEADERBOARD_FILE.exists() or not USERS_FILE.exists() or not QUIZZES_FILE.exists():
        return []
    
    lb_wb = load_workbook(LEADERBOARD_FILE)
    lb_ws = lb_wb.active
    
    users = {row[0]: row[1] for row in load_workbook(USERS_FILE).active.iter_rows(min_row=2, values_only=True)}
    quizzes = {row[0]: row[1] for row in load_workbook(QUIZZES_FILE).active.iter_rows(min_row=2, values_only=True)}
    
    results = []
    for row in lb_ws.iter_rows(min_row=2, values_only=True):
        results.append({
            "student_name": users.get(row[1], "Unknown"),
            "quiz_title": quizzes.get(row[2], "Deleted Quiz"),
            "score": row[3],
            "time_taken": row[4]
        })
    
    results.sort(key=lambda x: x['score'], reverse=True)
    return results

def get_all_teachers():
    if not USERS_FILE.exists():
        return []
    
    wb = load_workbook(USERS_FILE)
    ws = wb.active
    return [{"id": row[0], "username": row[1]} 
            for row in ws.iter_rows(min_row=2, values_only=True) 
            if row[3] == "teacher"]

def add_new_quiz(title, description=""):
    wb = load_workbook(QUIZZES_FILE)
    ws = wb.active
    
    # Get the last ID and increment by 1
    last_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            last_id = max(last_id, row[0])
    
    new_id = last_id + 1
    ws.append([new_id, title, description])
    wb.save(QUIZZES_FILE)
    return new_id

def delete_quiz(quiz_id):
    from openpyxl import load_workbook
    
    try:
        wb = load_workbook(QUIZZES_FILE)
        ws = wb.active
        
        for idx in range(ws.max_row, 1, -1):
            if ws.cell(row=idx, column=1).value == quiz_id:  # ID কলাম
                ws.delete_rows(idx)
                break
        
        wb.save(QUIZZES_FILE)

        if QUESTIONS_FILE.exists():
            wb_q = load_workbook(QUESTIONS_FILE)
            ws_q = wb_q.active
            
            for idx in range(ws_q.max_row, 1, -1):
                if ws_q.cell(row=idx, column=2).value == quiz_id:  # quiz_id কলাম
                    ws_q.delete_rows(idx)
            
            wb_q.save(QUESTIONS_FILE)
        
        return True
    except Exception as e:
        print(f"Error deleting quiz: {str(e)}")
        return False

def clean_student_data():
    if not USERS_FILE.exists():
        return
    
    wb = load_workbook(USERS_FILE)
    ws = wb.active
    
    seen = set()
    rows_to_delete = []
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 4:
            rows_to_delete.append(idx)
            continue
            
        username = row[1]
        if username in seen or row[3] != "student":
            rows_to_delete.append(idx)
        else:
            seen.add(username)

    for idx in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(idx)
    
    wb.save(USERS_FILE)