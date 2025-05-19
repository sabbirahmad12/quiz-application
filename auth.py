from excel_db import init_excel_db, load_workbook
import hashlib
from openpyxl import Workbook, load_workbook
import os
from pathlib import Path
from excel_db import (
    USERS_FILE
)

DATA_DIR = Path(__file__).parent / "data"
USERS_FILE = DATA_DIR / "users.xlsx"

def init_excel_db():
    if not DATA_DIR.exists():
        os.makedirs(DATA_DIR)
    
    if not USERS_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(["id", "username", "password", "role"])
        wb.save(USERS_FILE)

def login_user(username, password):
    try:
        hashed_pw = hashlib.sha256(password.encode()).hexdigest()

        init_excel_db()
        
        wb = load_workbook(USERS_FILE)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 4:
                if str(row[1]).strip() == username.strip() and str(row[2]).strip() == hashed_pw:
                    return {
                        'id': row[0],
                        'name': username,
                        'role': row[3]
                    }
        return None
    except Exception as e:
        print(f"Login error: {e}")
        return None


def register_user(username, password, role):
    try:
        # Password Hash
        hashed_pw = hashlib.sha256(password.encode()).hexdigest()
        
        init_excel_db()
        
        wb = load_workbook(USERS_FILE)
        ws = wb.active
        
        # Check Username
        for row in ws.iter_rows(values_only=True):
            if row and len(row) > 1 and row[1] == username:
                return False
                
        # Create New id
        existing_ids = [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row]
        new_id = max(existing_ids) + 1 if existing_ids else 1
        
        # Create New Username
        ws.append([new_id, username, hashed_pw, role])
        wb.save(USERS_FILE)
        return True
        
    except Exception as e:
        print(f"Registration error: {e}")
        return False